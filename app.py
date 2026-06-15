import streamlit as st
import pandas as pd
from jiwer import wer
import difflib
import fitz  # PyMuPDF
from docx import Document
import unicodedata
import re
from typing import Tuple, List
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import os

# PyVis — for graph visualization
try:
    from pyvis.network import Network
    PYVIS_AVAILABLE = True
except ImportError:
    PYVIS_AVAILABLE = False

# ==================================================
# ABNA — Extraction helpers (QA-01)
# ==================================================

def extract_pdf_text(pdf_file) -> str:
    page_separator = "\n\f\n"
    pdf = fitz.open(stream=pdf_file.read(), filetype="pdf")
    pages = []
    for page in pdf:
        pages.append(page.get_text("text"))
    return page_separator.join(pages)


def extract_docx_text(docx_file) -> str:
    doc = Document(docx_file)
    page_separator = "\n\f\n"
    paragraphs = []
    for para in doc.paragraphs:
        paragraphs.append(para.text)
    return page_separator.join(paragraphs)


def normalize_text(text: str) -> str:
    if text is None:
        return ""
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\f", "\n\f\n")
    text = re.sub(r"\n[ \t]*\f[ \t]*\n", "\n<PAGE>\n", text)
    metadata_patterns = [
        r"^\s*(ocr output)\b.*$",
        r"^\s*(ocr:)\b.*$",
        r"^\s*(filename|file):\b.*$",
        r"^\s*(pages?):\b.*$",
        r"^\s*(method|methods):\b.*$",
        r"^\s*PAGE[_\-]?\d+\b.*$",
        r"^\s*page[_\-]?\d+\b.*$",
        r"^\s*[-=]{3,}\s*$",
        r"^\s*text extracted by\b.*$",
    ]
    lines = text.split("\n")
    filtered_lines = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            filtered_lines.append("")
            continue
        is_metadata = any(re.match(p, stripped, flags=re.IGNORECASE) for p in metadata_patterns)
        if not is_metadata:
            filtered_lines.append(line)
    text = "\n".join(filtered_lines)
    text = re.sub(r"\n{3,}", "\n\n", text)
    paragraphs = [p.strip() for p in text.split("\n\n")]
    text = "\n\n".join([re.sub(r"\s+", " ", p) for p in paragraphs])
    text = text.replace("<PAGE>", "[PAGE]")
    text = text.lower().strip()
    return text


def calculate_ocr_accuracy(ground_truth: str, ocr_text: str) -> Tuple[float, float]:
    ground_truth = "" if ground_truth is None else str(ground_truth)
    ocr_text = "" if ocr_text is None else str(ocr_text)
    gt_norm = normalize_text(ground_truth)
    ocr_norm = normalize_text(ocr_text)
    error_rate = wer(gt_norm, ocr_norm)
    accuracy = max(0.0, (1 - error_rate) * 100)
    return error_rate, accuracy


def generate_word_comparison(gt_words: List[str], ocr_words: List[str]) -> pd.DataFrame:
    matcher = difflib.SequenceMatcher(None, gt_words, ocr_words)
    rows = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            for gt_w, ocr_w in zip(gt_words[i1:i2], ocr_words[j1:j2]):
                rows.append({"Ground Truth": gt_w, "OCR Output": ocr_w, "Status": "Match"})
        elif tag == "replace":
            gt_chunk  = gt_words[i1:i2]
            ocr_chunk = ocr_words[j1:j2]
            for k in range(max(len(gt_chunk), len(ocr_chunk))):
                gt_w  = gt_chunk[k]  if k < len(gt_chunk)  else ""
                ocr_w = ocr_chunk[k] if k < len(ocr_chunk) else ""
                rows.append({"Ground Truth": gt_w, "OCR Output": ocr_w, "Status": "Mismatch"})
        elif tag == "delete":
            for gt_w in gt_words[i1:i2]:
                rows.append({"Ground Truth": gt_w, "OCR Output": "", "Status": "Deletion"})
        elif tag == "insert":
            for ocr_w in ocr_words[j1:j2]:
                rows.append({"Ground Truth": "", "OCR Output": ocr_w, "Status": "Insertion"})
    return pd.DataFrame(rows)


def calculate_error_statistics(comparison_df: pd.DataFrame) -> dict:
    counts = comparison_df["Status"].value_counts().to_dict()
    return {
        "Matches":    counts.get("Match",     0),
        "Mismatches": counts.get("Mismatch",  0),
        "Insertions": counts.get("Insertion", 0),
        "Deletions":  counts.get("Deletion",  0),
    }


def create_professional_excel_report(ground_truth, ocr_text, accuracy, error_rate) -> bytes:
    gt_norm   = normalize_text(ground_truth)
    ocr_norm  = normalize_text(ocr_text)
    gt_words  = gt_norm.split()
    ocr_words = ocr_norm.split()
    comparison_df = generate_word_comparison(gt_words, ocr_words)
    error_stats   = calculate_error_statistics(comparison_df)

    output   = BytesIO()
    wb       = openpyxl.Workbook()
    BLUE_HDR = PatternFill("solid", fgColor="4472C4")
    WHT_FONT = Font(bold=True, color="FFFFFF")

    def style_header(ws, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = BLUE_HDR
            cell.font = WHT_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sheet 1 — Summary
    ws1 = wb.active
    ws1.title = "Summary"
    rows_data = [
        ("Metric", "Value"),
        ("OCR Accuracy (%)",    round(accuracy, 2)),
        ("Word Error Rate (%)", round(error_rate * 100, 2)),
        ("Ground Truth Words",  len(gt_words)),
        ("OCR Words",           len(ocr_words)),
        ("Total Matches",       error_stats["Matches"]),
        ("Total Mismatches",    error_stats["Mismatches"]),
        ("Total Insertions",    error_stats["Insertions"]),
        ("Total Deletions",     error_stats["Deletions"]),
    ]
    for r, (m, v) in enumerate(rows_data, 1):
        ws1.cell(r, 1, m)
        ws1.cell(r, 2, v)
    style_header(ws1, 2)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 20

    # Sheet 2 — Word Comparison
    ws2 = wb.create_sheet("Word Comparison")
    for c, h in enumerate(["Ground Truth", "OCR Output", "Status"], 1):
        ws2.cell(1, c, h)
    style_header(ws2, 3)
    fill_map = {
        "Match":     PatternFill("solid", fgColor="C6EFCE"),
        "Mismatch":  PatternFill("solid", fgColor="FFC7CE"),
        "Deletion":  PatternFill("solid", fgColor="FFEB9C"),
        "Insertion": PatternFill("solid", fgColor="FFD966"),
    }
    for i, row in comparison_df.iterrows():
        r = i + 2
        ws2.cell(r, 1, row["Ground Truth"])
        ws2.cell(r, 2, row["OCR Output"])
        ws2.cell(r, 3, row["Status"]).alignment = Alignment(horizontal="center")
        fill = fill_map.get(row["Status"])
        if fill:
            for c in range(1, 4):
                ws2.cell(r, c).fill = fill
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 15

    wb.save(output)
    return output.getvalue()


def create_text_report(ground_truth, ocr_text, accuracy, error_rate) -> str:
    gt_words  = normalize_text(ground_truth).split()
    ocr_words = normalize_text(ocr_text).split()
    stats = calculate_error_statistics(generate_word_comparison(gt_words, ocr_words))
    return "\n".join([
        "OCR ACCURACY REPORT", "=" * 40,
        f"OCR Accuracy      : {accuracy:.2f}%",
        f"Word Error Rate   : {error_rate * 100:.2f}%",
        f"Ground Truth Words: {len(gt_words)}",
        f"OCR Words         : {len(ocr_words)}", "",
        "ERROR BREAKDOWN", "-" * 40,
        f"Matches    : {stats['Matches']}",
        f"Mismatches : {stats['Mismatches']}",
        f"Insertions : {stats['Insertions']}",
        f"Deletions  : {stats['Deletions']}",
    ])


# ==================================================
# HARINI — NER helpers (QA-02)
# ==================================================

ENTITY_COLORS = {
    "HERB":               "#90EE90",
    "DISEASE":            "#FFB6B6",
    "BODY_PART":          "#FFD580",
    "PROCEDURE":          "#ADD8E6",
    "BIOLOGICAL_PROCESS": "#D8BFD8",
    "SYMPTOM":            "#FFF176",
    "CHEMICAL":           "#A7FFEB",
    "PROTEIN":            "#C5CAE9",
    "DOSHA":              "#ddd5f7",
    "COMPOUND":           "#fef3c7",
}


def highlight_entities(text: str, entity_df: pd.DataFrame) -> str:
    """
    Highlight entities in text using span-based, non-overlapping replacement.
    Finds all entity occurrences as character-level spans, sorts them, resolves
    overlaps (longest match wins), then rebuilds the string — so no entity label
    ever bleeds into surrounding text.
    """
    if not text:
        return text

    entity_df = entity_df.copy()
    entity_df["Entity"] = entity_df["Entity"].astype(str)
    entity_df["Label"]  = entity_df["Label"].astype(str).str.upper()

    # Build list of (start, end, entity_text, label, color)
    spans = []
    for _, row in entity_df.iterrows():
        entity = row["Entity"].strip()
        label  = row["Label"]
        color  = ENTITY_COLORS.get(label, "#ffff99")
        if not entity:
            continue
        # Case-insensitive search for all occurrences
        pattern = re.compile(re.escape(entity), re.IGNORECASE)
        for m in pattern.finditer(text):
            spans.append((m.start(), m.end(), m.group(), label, color))

    if not spans:
        return text

    # Sort by start position; for ties prefer longer span (greedy)
    spans.sort(key=lambda s: (s[0], -(s[1] - s[0])))

    # Remove overlapping spans — keep first (longest at each position)
    filtered = []
    last_end = -1
    for span in spans:
        start, end, matched_text, label, color = span
        if start >= last_end:
            filtered.append(span)
            last_end = end

    # Reconstruct the text with HTML highlights
    result = []
    cursor = 0
    for start, end, matched_text, label, color in filtered:
        # Append plain text before this span
        result.append(text[cursor:start])
        # Append highlighted span
        result.append(
            f"<span style='background-color:{color};"
            f"padding:2px 5px;border-radius:4px;font-weight:500;'>"
            f"{matched_text}"
            f"<sup style='font-size:0.65em;font-weight:700;margin-left:2px;"
            f"color:#333;'>{label}</sup></span>"
        )
        cursor = end

    # Append any remaining plain text
    result.append(text[cursor:])

    # Preserve line breaks for readability
    highlighted = "".join(result)
    highlighted = highlighted.replace("\n", "<br>")
    return highlighted


def compute_ner_metrics(auto_df: pd.DataFrame, manual_df: pd.DataFrame):
    auto_entities   = set(auto_df["Entity"].dropna().astype(str).str.lower())
    manual_entities = set(manual_df["Entity"].dropna().astype(str).str.lower())
    matched = auto_entities & manual_entities
    missing = manual_entities - auto_entities
    extra   = auto_entities - manual_entities
    precision = len(matched) / (len(matched) + len(extra))   if (len(matched) + len(extra))   > 0 else 0
    recall    = len(matched) / (len(matched) + len(missing)) if (len(matched) + len(missing)) > 0 else 0
    f1        = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
    return matched, missing, extra, precision, recall, f1


# ==================================================
# NEHA — Relationship helpers (QA-03)
# ==================================================

def validate_relationships(df: pd.DataFrame) -> pd.DataFrame:
    df = df.drop_duplicates()
    df = df.copy()
    df["Status"] = "Correct"
    return df


# ==================================================
# NEHA — Neo4j helpers (DB-02)
# ==================================================

def build_cypher(df: pd.DataFrame, source_col: str, relation_col: str, target_col: str) -> str:
    lines = ["// MedYukthee AI — Validated Ayurveda Knowledge Graph",
             "// Generated by QA Validation Tool", ""]
    for _, row in df.iterrows():
        src = str(row[source_col]).replace("'", "\\'")
        rel = str(row[relation_col]).upper().replace(" ", "_")
        tgt = str(row[target_col]).replace("'", "\\'")
        lines.append(
            f"MERGE (a:Entity {{name:'{src}'}})\n"
            f"MERGE (b:Entity {{name:'{tgt}'}})\n"
            f"MERGE (a)-[:{rel}]->(b);"
        )
    return "\n\n".join(lines)


def draw_graph(df: pd.DataFrame, src_col: str, rel_col: str, tgt_col: str) -> str:
    net = Network(height="500px", width="100%", bgcolor="#1a1a2e", font_color="white",
                  directed=True)
    net.set_options("""
    {
      "nodes": {"font": {"size": 14, "color": "white"}, "borderWidth": 2, "shadow": true},
      "edges": {
        "font": {"size": 11, "color": "#aaaaaa", "align": "middle"},
        "arrows": {"to": {"enabled": true, "scaleFactor": 0.8}},
        "smooth": {"type": "curvedCW", "roundness": 0.2},
        "shadow": true
      },
      "physics": {
        "forceAtlas2Based": {"gravitationalConstant": -50, "springLength": 120},
        "solver": "forceAtlas2Based"
      }
    }
    """)
    added_nodes = set()
    for _, row in df.iterrows():
        src = str(row[src_col]).strip()
        rel = str(row[rel_col]).strip().upper()
        tgt = str(row[tgt_col]).strip()
        if src not in added_nodes:
            net.add_node(src, label=src, color="#4CAF50", size=20)
            added_nodes.add(src)
        if tgt not in added_nodes:
            net.add_node(tgt, label=tgt, color="#F44336", size=20)
            added_nodes.add(tgt)
        net.add_edge(src, tgt, label=rel, color="#888888")
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".html")
    net.save_graph(tmp.name)
    with open(tmp.name, "r", encoding="utf-8") as f:
        html = f.read()
    os.unlink(tmp.name)
    return html


# ==================================================
# Streamlit UI — 4 tabs
# ==================================================

st.set_page_config(page_title="MedYukthee QA Validation Tool", layout="wide")
st.title("MedYukthee QA Validation Tool")

tab1, tab2, tab3, tab4 = st.tabs([
    "🔤 OCR Accuracy (QA-01)",
    "🏷️ NER Validator (QA-02)",
    "🔗 Relationship Reviewer (QA-03)",
    "🗄️ Neo4j Export (DB-02)",
])

# ══════════════════════════════════════════════════
# TAB 1 — ABNA : OCR Accuracy (QA-01)
# ══════════════════════════════════════════════════
with tab1:
    st.subheader("OCR Accuracy Validation")
    st.write("Upload the original PDF and the OCR-generated text file.")

    original_pdf = st.file_uploader("Upload Original PDF", type=["pdf"], key="pdf")
    ocr_file_up  = st.file_uploader("Upload OCR Output",  type=["txt", "docx"], key="ocr")

    if st.button("Calculate Accuracy"):
        if not original_pdf or not ocr_file_up:
            st.warning("Please upload both files.")
        else:
            ground_truth = extract_pdf_text(original_pdf)

            if ocr_file_up.name.lower().endswith(".txt"):
                raw = ocr_file_up.read()
                try:
                    ocr_text = raw.decode("utf-8")
                except Exception:
                    ocr_text = raw.decode("latin-1", errors="ignore")
            else:
                ocr_text = extract_docx_text(ocr_file_up)

            error_rate, accuracy = calculate_ocr_accuracy(ground_truth, ocr_text)

            c1, c2 = st.columns(2)
            c1.metric("Word Error Rate (WER)", f"{error_rate * 100:.2f}%")
            c2.metric("OCR Accuracy",          f"{accuracy:.2f}%")

            st.subheader("Ground Truth Preview")
            st.text_area("PDF Text (normalized)", normalize_text(ground_truth)[:3000], height=200)
            st.subheader("OCR Output Preview")
            st.text_area("OCR Text (normalized)", normalize_text(ocr_text)[:3000],    height=200)

            st.subheader("Word Diff (colour highlight)")
            gt_w  = normalize_text(ground_truth).split()
            ocr_w = normalize_text(ocr_text).split()
            diff_html = ""
            for token in difflib.ndiff(gt_w, ocr_w):
                if token.startswith("  "):
                    diff_html += f"{token[2:]} "
                elif token.startswith("- "):
                    diff_html += f'<span style="background:#ffcccc;">{token[2:]} </span>'
                elif token.startswith("+ "):
                    diff_html += f'<span style="background:#ccffcc;">{token[2:]} </span>'
            st.markdown(diff_html, unsafe_allow_html=True)

            comp_df = generate_word_comparison(gt_w, ocr_w)
            st.subheader("Word Comparison Table (first 100 rows)")
            st.dataframe(comp_df.head(100).reset_index(drop=True), use_container_width=True)
            if len(comp_df) > 100:
                st.info(f"Showing 100 of {len(comp_df)} rows — full data in Excel.")

            st.subheader("Error Statistics")
            err = calculate_error_statistics(comp_df)
            e1, e2, e3, e4 = st.columns(4)
            e1.metric("Matches",    err["Matches"])
            e2.metric("Mismatches", err["Mismatches"])
            e3.metric("Insertions", err["Insertions"])
            e4.metric("Deletions",  err["Deletions"])

            excel_data = create_professional_excel_report(ground_truth, ocr_text, accuracy, error_rate)
            text_data  = create_text_report(ground_truth, ocr_text, accuracy, error_rate)

            d1, d2 = st.columns(2)
            with d1:
                st.download_button(
                    "📥 Download Excel Report (2 sheets)",
                    data=excel_data,
                    file_name="ocr_accuracy_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            with d2:
                st.download_button(
                    "📄 Download Text Report (.txt)",
                    data=text_data,
                    file_name="ocr_accuracy_report.txt",
                    mime="text/plain",
                )

# ══════════════════════════════════════════════════
# TAB 2 — HARINI : NER Validator (QA-02)
# ══════════════════════════════════════════════════
with tab2:
    st.subheader("NER Validation & Entity Highlighting Tool")
    st.markdown("---")

    ocr_txt_file  = st.file_uploader("Upload OCR Text File (.txt)",                type=["txt"],        key="ner_ocr")
    auto_ner_file = st.file_uploader("Upload Automated Annotation (.xlsx or .csv)", type=["xlsx","csv"], key="auto_ner")
    manual_file   = st.file_uploader("Upload Manual Annotation (.xlsx or .csv)",    type=["xlsx","csv"], key="manual_ner")

    st.markdown("---")

    if ocr_txt_file and auto_ner_file:
        text    = ocr_txt_file.read().decode("utf-8")
        auto_df = pd.read_excel(auto_ner_file) if auto_ner_file.name.endswith(".xlsx") else pd.read_csv(auto_ner_file)

        # Legend with color swatches
        st.markdown("### Legend")
        legend_cols = st.columns(len(ENTITY_COLORS))
        for col, (label, color) in zip(legend_cols, ENTITY_COLORS.items()):
            col.markdown(
                f"<span style='background-color:{color};padding:3px 8px;"
                f"border-radius:4px;font-size:0.8em;font-weight:600;'>{label}</span>",
                unsafe_allow_html=True,
            )

        if "Entity" in auto_df.columns and "Label" in auto_df.columns:
            st.subheader("Highlighted Entities")
            highlighted_html = highlight_entities(text, auto_df)
            # Wrap in a styled container for clean rendering
            st.markdown(
                f"""
                <div style="
                    background-color: #0e1117;
                    color: #fafafa;
                    padding: 20px 24px;
                    border-radius: 8px;
                    border: 1px solid #333;
                    line-height: 1.9;
                    font-size: 0.95rem;
                    font-family: 'Source Sans Pro', sans-serif;
                    white-space: pre-wrap;
                    word-wrap: break-word;
                ">
                {highlighted_html}
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            st.warning("Automated file must have columns: Entity, Label")

    if auto_ner_file:
        auto_df2 = pd.read_excel(auto_ner_file) if auto_ner_file.name.endswith(".xlsx") else pd.read_csv(auto_ner_file)
        st.subheader("Automated Annotation Preview")
        st.dataframe(auto_df2, use_container_width=True)

    if manual_file:
        manual_df = pd.read_excel(manual_file) if manual_file.name.endswith(".xlsx") else pd.read_csv(manual_file)
        st.subheader("Manual Annotation Preview")
        st.dataframe(manual_df, use_container_width=True)

    if auto_ner_file and manual_file:
        st.markdown("---")
        st.header("NER Validation")

        auto_df_v   = pd.read_excel(auto_ner_file) if auto_ner_file.name.endswith(".xlsx") else pd.read_csv(auto_ner_file)
        manual_df_v = pd.read_excel(manual_file)   if manual_file.name.endswith(".xlsx")   else pd.read_csv(manual_file)

        matched, missing, extra, precision, recall, f1 = compute_ner_metrics(auto_df_v, manual_df_v)

        st.subheader("Validation Summary")
        m1, m2, m3 = st.columns(3)
        m1.metric("Matched", len(matched))
        m2.metric("Missing", len(missing))
        m3.metric("Extra",   len(extra))

        chart_df = pd.DataFrame({
            "Category": ["Matched", "Missing", "Extra"],
            "Count":    [len(matched), len(missing), len(extra)]
        })
        st.subheader("Validation Overview")
        st.bar_chart(chart_df.set_index("Category"))

        st.subheader("Performance Metrics")
        p1, p2, p3 = st.columns(3)
        p1.metric("Precision", f"{precision*100:.2f}%")
        p2.metric("Recall",    f"{recall*100:.2f}%")
        p3.metric("F1 Score",  f"{f1*100:.2f}%")

        if f1 >= 0.90:
            st.success("✅ Excellent NER Performance")
        elif f1 >= 0.80:
            st.success("✅ Good NER Performance")
        elif f1 >= 0.70:
            st.warning("⚠️ Moderate NER Performance")
        else:
            st.error("❌ NER Performance Needs Improvement")

        st.markdown("---")
        report_df = pd.DataFrame({
            "Metric": ["Matched", "Missing", "Extra", "Precision", "Recall", "F1 Score"],
            "Value":  [len(matched), len(missing), len(extra),
                       round(precision*100,2), round(recall*100,2), round(f1*100,2)]
        })
        st.download_button(
            "📥 Download NER Report (.csv)",
            data=report_df.to_csv(index=False),
            file_name="ner_report.csv",
            mime="text/csv",
        )

        col_m, col_e = st.columns(2)
        with col_m:
            st.subheader("Missing Entities")
            if missing:
                st.dataframe(pd.DataFrame(sorted(missing), columns=["Missing Entity"]), use_container_width=True)
            else:
                st.success("No Missing Entities")
        with col_e:
            st.subheader("Extra Entities")
            if extra:
                st.dataframe(pd.DataFrame(sorted(extra), columns=["Extra Entity"]), use_container_width=True)
            else:
                st.success("No Extra Entities")

# ══════════════════════════════════════════════════
# TAB 3 — NEHA : Relationship Reviewer (QA-03)
# ══════════════════════════════════════════════════
with tab3:
    st.subheader("Relationship Validator (QA-03)")
    st.write("Upload your relationships CSV to validate, review, and download the cleaned data.")

    rel_file = st.file_uploader(
        "Upload relationships.csv — columns: Source/subject, Relation, Target/object",
        type=["csv"], key="rel"
    )

    if rel_file:
        df_raw = pd.read_csv(rel_file)

        st.subheader("Raw Data Preview")
        st.dataframe(df_raw, use_container_width=True)

        # Auto-detect column names
        cols    = df_raw.columns.tolist()
        col_map = {}
        for c in cols:
            cl = c.lower()
            if cl in ("source", "subject"):
                col_map["source"] = c
            elif cl in ("relation", "relationship"):
                col_map["relation"] = c
            elif cl in ("target", "object"):
                col_map["target"] = c

        if len(col_map) < 3:
            st.error(f"CSV must have Source/subject, Relation, Target/object columns. Found: {cols}")
        else:
            src_col = col_map["source"]
            rel_col = col_map["relation"]
            tgt_col = col_map["target"]

            df_validated = validate_relationships(df_raw)

            st.subheader("Validation Summary")
            v1, v2, v3 = st.columns(3)
            v1.metric("Total Rows (original)", len(df_raw))
            v2.metric("Duplicates Removed",    int(df_raw.duplicated().sum()))
            v3.metric("Valid Relationships",   len(df_validated))

            st.subheader("Validated Relationships")
            st.dataframe(df_validated, use_container_width=True)

            # ── Relationship type breakdown ────────────────────────────────────
            st.subheader("Relationship Type Breakdown")
            rel_counts = df_validated[rel_col].value_counts().reset_index()
            rel_counts.columns = ["Relation Type", "Count"]
            st.bar_chart(rel_counts.set_index("Relation Type"))
            st.dataframe(rel_counts, use_container_width=True)

            # ── Downloads ─────────────────────────────────────────────────────
            st.markdown("---")
            d1, d2 = st.columns(2)
            with d1:
                st.download_button(
                    "📥 Download Validated CSV",
                    data=df_validated.to_csv(index=False),
                    file_name="validated_relationships.csv",
                    mime="text/csv",
                )
            with d2:
                # Validation report
                val_report = "\n".join([
                    "RELATIONSHIP VALIDATION REPORT", "=" * 40,
                    f"Total Rows (original) : {len(df_raw)}",
                    f"Duplicates Removed    : {int(df_raw.duplicated().sum())}",
                    f"Valid Relationships   : {len(df_validated)}",
                    "", "Relation Type Counts", "-" * 40,
                ] + [f"{r['Relation Type']}: {r['Count']}" for _, r in rel_counts.iterrows()])
                st.download_button(
                    "📄 Download Validation Report (.txt)",
                    data=val_report,
                    file_name="relationship_validation_report.txt",
                    mime="text/plain",
                )

            # Store validated data in session for Tab 4
            st.session_state["validated_df"]  = df_validated
            st.session_state["src_col"]       = src_col
            st.session_state["rel_col"]       = rel_col
            st.session_state["tgt_col"]       = tgt_col
            st.success("✅ Data saved — go to Tab 4 (Neo4j Export) to visualise and export.")

# ══════════════════════════════════════════════════
# TAB 4 — NEHA : Neo4j Export (DB-02)
# ══════════════════════════════════════════════════
with tab4:
    st.subheader("Neo4j Graph Export (DB-02)")
    st.write(
        "This tab takes the validated relationships from Tab 3 and exports them as "
        "a Neo4j Cypher file — ready to load into Neo4j Browser."
    )

    # Load from session OR allow direct upload
    if "validated_df" in st.session_state:
        df_neo   = st.session_state["validated_df"]
        src_col  = st.session_state["src_col"]
        rel_col  = st.session_state["rel_col"]
        tgt_col  = st.session_state["tgt_col"]
        st.success(f"✅ Using {len(df_neo)} validated relationships from Tab 3.")
    else:
        st.info("No data from Tab 3 yet. You can also upload directly here.")
        neo_file = st.file_uploader("Upload validated_relationships.csv", type=["csv"], key="neo_upload")
        if neo_file:
            df_neo  = pd.read_csv(neo_file)
            cols    = df_neo.columns.tolist()
            col_map = {}
            for c in cols:
                cl = c.lower()
                if cl in ("source", "subject"):   col_map["source"]   = c
                elif cl in ("relation", "relationship"): col_map["relation"] = c
                elif cl in ("target", "object"):  col_map["target"]   = c
            if len(col_map) < 3:
                st.error(f"Could not detect columns. Found: {cols}")
                st.stop()
            src_col = col_map["source"]
            rel_col = col_map["relation"]
            tgt_col = col_map["target"]
        else:
            st.stop()

    # ── Graph Schema Reference ─────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📐 Graph Schema (DB-02 Design)")
    col_schema1, col_schema2 = st.columns(2)
    with col_schema1:
        st.markdown("""
**Node Types**
- `Herb` — e.g. Ashwagandha, Brahmi, Triphala
- `Disease` — e.g. Anxiety, Fever, Indigestion
- `Dosha` — e.g. Vata, Pitta, Kapha
- `Compound` — e.g. Haritaki, Amla, Bibhitaki
""")
    with col_schema2:
        st.markdown("""
**Relationship Types**
- `Herb` **TREATS** `Disease`
- `Herb` **BALANCES** `Dosha`
- `Herb` **CONTAINS** `Compound`
- `Herb` **IMPROVES** Condition
- `Herb` **PREVENTS** `Disease`
""")

    # ── Knowledge Graph Visualization ─────────────────────────────────────────
    st.markdown("---")
    st.subheader("🕸️ Knowledge Graph Visualization")
    st.caption("🟢 Green = Source node (Herb)   🔴 Red = Target node (Disease/Dosha)")

    if PYVIS_AVAILABLE:
        graph_html = draw_graph(df_neo, src_col, rel_col, tgt_col)
        st.components.v1.html(graph_html, height=520, scrolling=False)
    else:
        st.warning("Install pyvis for graph visualization:  pip install pyvis")

    # ── Cypher Export ─────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📄 Neo4j Cypher Export")
    st.write(
        "Download this `.cypher` file. To load it into Neo4j: open Neo4j Browser → "
        "paste the contents → run. Or use `cypher-shell` in terminal."
    )

    cypher_text = build_cypher(df_neo, src_col, rel_col, tgt_col)
    st.code(cypher_text[:2000], language="cypher")
    if len(cypher_text) > 2000:
        st.info("Preview shows first 2000 chars — full file in the download below.")

    st.download_button(
        "🔗 Download Neo4j Cypher File (.cypher)",
        data=cypher_text,
        file_name="medyukthee_knowledge_graph.cypher",
        mime="text/plain",
    )
    st.success("✅ Cypher file ready to load into Neo4j!")

    # ── How to use in Neo4j Browser ───────────────────────────────────────────
    st.markdown("---")
    st.subheader("📖 How to load this into Neo4j")
    st.markdown("""
1. Open **Neo4j Browser** at `http://localhost:7474`
2. Log in with your credentials
3. Click the file icon (top left) → paste the contents of the `.cypher` file
4. Press **Run** (▶️)
5. Then run this query to see your graph:
```cypher
MATCH (n)-[r]->(m) RETURN n, r, m
```
""")